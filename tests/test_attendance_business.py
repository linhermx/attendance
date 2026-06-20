from __future__ import annotations

from datetime import date, datetime
from pathlib import Path
from tempfile import TemporaryDirectory
import unittest

from openpyxl import load_workbook
import pandas as pd

from attendance.classification import clasificar_checadas, load_classification_configuration
from attendance.core import (
    analyze_operational_day,
    build_quick_view_source,
    calculate_attendance,
    calculate_attendance_range,
    calculate_worked_minutes,
    schedule_for_date,
)


WEEKDAY_SHIFT = {
    "entrada": "08:00:00",
    "inicio_comida": "12:00:00",
    "fin_comida": "12:45:00",
    "salida": "17:00:00",
}
SATURDAY_SHIFT = {
    "entrada": "08:00:00",
    "inicio_comida": "12:00:00",
    "fin_comida": "12:30:00",
    "salida": "14:00:00",
}
REMOVED_VISIBLE_LABEL = "Horas" + " extra"


def prepared_personal() -> pd.DataFrame:
    return pd.DataFrame(
        [
            {"id_usuario": "1", "nombre_completo": "PERSONA UNO"},
            {"id_usuario": "2", "nombre_completo": "PERSONA DOS"},
        ]
    )


def prepared_events(work_date: date, values: dict[str, tuple[str, ...]]) -> pd.DataFrame:
    return pd.DataFrame(
        [
            {
                "id_usuario": employee_id,
                "tiempo": pd.Timestamp(f"{work_date} {checked_at}"),
                "estado_normalizado": "",
            }
            for employee_id, times in values.items()
            for checked_at in times
        ],
        columns=["id_usuario", "tiempo", "estado_normalizado"],
    )


def operational_row(
    times: list[str],
    *,
    work_date: date = date(2026, 6, 5),
    cutoff_time: datetime | None = None,
) -> pd.Series:
    personal = prepared_personal().iloc[[0]].copy()
    events = prepared_events(work_date, {"1": tuple(times)})
    return analyze_operational_day(
        personal,
        events,
        work_date,
        schedule_for_date(work_date, []),
        cutoff_time=cutoff_time,
    ).iloc[0]


def write_sources(
    root: Path,
    events_by_employee: dict[str, tuple[str, str, tuple[str, ...]]],
    *,
    work_date: date,
) -> tuple[Path, Path]:
    personal_path = root / "personal.xlsx"
    events_path = root / "events.xlsx"
    pd.DataFrame(
        [
            {"ID de usuario": employee_id, "Nombre": name, "Apellido": last_name}
            for employee_id, (name, last_name, _times) in events_by_employee.items()
        ]
    ).to_excel(personal_path, sheet_name="data", index=False)
    pd.DataFrame(
        [
            {
                "Tiempo": f"{work_date} {checked_at}",
                "ID de usuario": employee_id,
                "Nombre": name,
                "Apellido": last_name,
                "Estado": "",
            }
            for employee_id, (name, last_name, times) in events_by_employee.items()
            for checked_at in times
        ]
    ).to_excel(events_path, sheet_name="data", index=False)
    return personal_path, events_path


class FlexibleLunchTests(unittest.TestCase):
    def test_saturday_lunch_zone_punch_is_lunch_out_not_final_exit(self) -> None:
        result = clasificar_checadas(["08:03:06", "12:03:24"], SATURDAY_SHIFT)
        self.assertEqual(result["entrada"], "08:03:06")
        self.assertEqual(result["inicio_comida"], "12:03:24")
        self.assertIsNone(result["fin_comida"])
        self.assertIsNone(result["salida"])
        self.assertEqual(result["status"], "Retardo + incidencia")
        self.assertEqual(
            result["detalle"],
            "Retardo (3 min) | Sin regreso de comida | Sin salida final",
        )
        self.assertNotIn("Salida anticipada", result["detalle"])

    def test_saturday_lunch_zone_with_final_exit_keeps_missing_lunch_return_only(self) -> None:
        result = clasificar_checadas(
            ["08:00:00", "12:05:00", "14:00:00"],
            SATURDAY_SHIFT,
        )
        self.assertEqual(result["entrada"], "08:00:00")
        self.assertEqual(result["inicio_comida"], "12:05:00")
        self.assertIsNone(result["fin_comida"])
        self.assertEqual(result["salida"], "14:00:00")
        self.assertEqual(result["detalle"], "Sin regreso de comida")
        self.assertNotIn("Salida anticipada", result["detalle"])

    def test_saturday_reference_lunch_sequence_is_punctual(self) -> None:
        result = clasificar_checadas(
            ["08:00:00", "12:00:00", "12:30:00", "14:00:00"],
            SATURDAY_SHIFT,
        )
        self.assertEqual(result["status"], "Puntual")
        self.assertEqual(result["detalle"], "")

    def test_weekday_lunch_zone_punch_is_lunch_out_not_final_exit(self) -> None:
        result = clasificar_checadas(["08:00:00", "12:05:00"], WEEKDAY_SHIFT)
        self.assertEqual(result["entrada"], "08:00:00")
        self.assertEqual(result["inicio_comida"], "12:05:00")
        self.assertIsNone(result["fin_comida"])
        self.assertIsNone(result["salida"])
        self.assertEqual(result["detalle"], "Sin regreso de comida | Sin salida final")
        self.assertNotIn("Salida anticipada", result["detalle"])

    def test_weekday_lunch_zone_with_final_exit_keeps_missing_lunch_return_only(self) -> None:
        result = clasificar_checadas(
            ["08:00:00", "12:05:00", "17:00:00"],
            WEEKDAY_SHIFT,
        )
        self.assertEqual(result["entrada"], "08:00:00")
        self.assertEqual(result["inicio_comida"], "12:05:00")
        self.assertIsNone(result["fin_comida"])
        self.assertEqual(result["salida"], "17:00:00")
        self.assertEqual(result["detalle"], "Sin regreso de comida")

    def test_weekday_protected_zone_complete_lunch_is_punctual(self) -> None:
        result = clasificar_checadas(
            ["08:00:00", "12:05:00", "12:50:00", "17:00:00"],
            WEEKDAY_SHIFT,
        )
        self.assertEqual(result["entrada"], "08:00:00")
        self.assertEqual(result["inicio_comida"], "12:05:00")
        self.assertEqual(result["fin_comida"], "12:50:00")
        self.assertEqual(result["salida"], "17:00:00")
        self.assertEqual(result["status"], "Puntual")
        self.assertEqual(result["detalle"], "")

    def test_weekday_protected_zone_lunch_excess_is_not_exit_or_ambiguous(self) -> None:
        result = clasificar_checadas(
            ["08:00:00", "12:05:00", "12:55:01", "17:00:00"],
            WEEKDAY_SHIFT,
        )
        self.assertEqual(result["inicio_comida"], "12:05:00")
        self.assertEqual(result["fin_comida"], "12:55:01")
        self.assertEqual(result["salida"], "17:00:00")
        self.assertEqual(result["status"], "Incidencia")
        self.assertEqual(result["detalle"], "Exceso de comida (+6 min)")
        self.assertNotIn("Salida anticipada", result["detalle"])
        self.assertNotIn("Registro ambiguo", result["detalle"])

    def test_weekday_flexible_45_minute_pairs_are_valid(self) -> None:
        for lunch_out, lunch_return in (
            ("12:00:00", "12:45:00"),
            ("12:20:00", "13:05:00"),
            ("12:40:00", "13:25:00"),
            ("13:10:00", "13:55:00"),
            ("14:00:00", "14:45:00"),
            ("15:00:00", "15:45:00"),
        ):
            with self.subTest(lunch_out=lunch_out):
                result = clasificar_checadas(
                    ["08:00:00", lunch_out, lunch_return, "17:00:00"],
                    WEEKDAY_SHIFT,
                )
                self.assertEqual(result["detalle"], "")
                self.assertEqual(result["inicio_comida"], lunch_out)
                self.assertEqual(result["fin_comida"], lunch_return)

    def test_weekday_45_00_is_valid_and_45_01_is_excess(self) -> None:
        exact = clasificar_checadas(
            ["08:00:00", "12:40:00", "13:25:00", "17:00:00"],
            WEEKDAY_SHIFT,
        )
        over = clasificar_checadas(
            ["08:00:00", "12:40:00", "13:25:01", "17:00:00"],
            WEEKDAY_SHIFT,
        )
        self.assertEqual(exact["detalle"], "")
        self.assertEqual(over["detalle"], "Exceso de comida (+1 min)")

    def test_saturday_30_00_is_valid_and_30_01_is_excess(self) -> None:
        exact = clasificar_checadas(
            ["08:00:00", "12:40:00", "13:10:00", "14:00:00"],
            SATURDAY_SHIFT,
        )
        over = clasificar_checadas(
            ["08:00:00", "12:40:00", "13:10:01", "14:00:00"],
            SATURDAY_SHIFT,
        )
        self.assertEqual(exact["detalle"], "")
        self.assertEqual(over["detalle"], "Exceso de comida (+1 min)")

    def test_50_minute_weekday_lunch_reports_only_five_minute_excess(self) -> None:
        result = clasificar_checadas(
            ["08:00:00", "13:10:00", "14:00:00", "17:00:00"],
            WEEKDAY_SHIFT,
        )
        self.assertEqual(result["detalle"], "Exceso de comida (+5 min)")

    def test_input_order_does_not_define_assignments(self) -> None:
        result = clasificar_checadas(
            ["17:00:00", "12:45:00", "08:00:00", "12:00:00"],
            WEEKDAY_SHIFT,
        )
        self.assertEqual(result["entrada"], "08:00:00")
        self.assertEqual(result["inicio_comida"], "12:00:00")
        self.assertEqual(result["fin_comida"], "12:45:00")
        self.assertEqual(result["salida"], "17:00:00")


class ContextualClassificationTests(unittest.TestCase):
    def test_explicit_device_states_are_used_without_making_lunch_rigid(self) -> None:
        result = clasificar_checadas(
            ["11:59:00", "12:09:00", "12:48:00"],
            WEEKDAY_SHIFT,
            estados=["Entrada", "Salida a descanso", "Regreso descanso"],
            dispositivos=["CHECADOR_PRODUCCION", "CHECADOR_PRODUCCION", "CHECADOR_PRODUCCION"],
        )
        self.assertEqual(result["entrada"], "11:59:00")
        self.assertEqual(result["inicio_comida"], "12:09:00")
        self.assertEqual(result["fin_comida"], "12:48:00")
        self.assertIsNone(result["salida"])
        self.assertEqual(result["detalle"], "Retardo grave (239 min) | Sin salida final")
        self.assertIn("estado=Salida a descanso", result["auditoria"])
        self.assertIn("dispositivo=CHECADOR_PRODUCCION", result["auditoria"])

    def test_blank_states_still_follow_context(self) -> None:
        result = clasificar_checadas(
            ["08:00:00", "12:00:00", "12:45:00", "17:00:00"],
            WEEKDAY_SHIFT,
            estados=["", "", "", ""],
            dispositivos=["CHECADOR_PRODUCCION"] * 4,
        )
        self.assertEqual(result["entrada"], "08:00:00")
        self.assertEqual(result["inicio_comida"], "12:00:00")
        self.assertEqual(result["fin_comida"], "12:45:00")
        self.assertEqual(result["salida"], "17:00:00")
        self.assertEqual(result["status"], "Puntual")
        self.assertEqual(result["detalle"], "")

    def test_invalid_states_fall_back_to_context(self) -> None:
        result = clasificar_checadas(
            ["08:00:00", "12:05:00", "12:45:00", "17:00:00"],
            WEEKDAY_SHIFT,
            estados=["Entrada", "Estado invalido", "Otro estado", "Salida"],
            dispositivos=["CHECADOR_PRODUCCION"] * 4,
        )
        self.assertEqual(result["entrada"], "08:00:00")
        self.assertEqual(result["inicio_comida"], "12:05:00")
        self.assertEqual(result["fin_comida"], "12:45:00")
        self.assertEqual(result["salida"], "17:00:00")
        self.assertEqual(result["status"], "Puntual")
        self.assertEqual(result["detalle"], "")
        self.assertNotIn("Salida anticipada", result["detalle"])

    def test_severe_tardy_with_complete_lunch_has_only_tardy_detail(self) -> None:
        result = clasificar_checadas(
            ["10:00:00", "13:00:00", "13:30:00", "17:00:00"],
            WEEKDAY_SHIFT,
        )
        self.assertEqual(result["status"], "Retardo")
        self.assertEqual(result["detalle"], "Retardo grave (120 min)")

    def test_isolated_lunch_time_punch_does_not_receive_an_invented_entry(self) -> None:
        result = clasificar_checadas(["12:01:00"], WEEKDAY_SHIFT)
        self.assertIsNone(result["entrada"])
        self.assertEqual(result["inicio_comida"], "12:01:00")
        self.assertEqual(
            result["detalle"],
            "Sin entrada | Sin regreso de comida | Sin salida final",
        )

    def test_isolated_lunch_return_reference_punch_is_not_left_blank(self) -> None:
        result = clasificar_checadas(
            ["12:44:56"],
            WEEKDAY_SHIFT,
            estados=["Estado invalido"],
            dispositivos=["CHECADOR_PRODUCCION"],
        )
        self.assertIsNone(result["entrada"])
        self.assertIsNone(result["inicio_comida"])
        self.assertEqual(result["fin_comida"], "12:44:56")
        self.assertIsNone(result["salida"])
        self.assertEqual(result["status"], "Incidencia")
        self.assertEqual(
            result["detalle"],
            "Sin entrada | Sin inicio de comida | Sin salida final",
        )
        self.assertNotIn("Registro ambiguo", result["detalle"])
        self.assertNotIn("Sin regreso de comida", result["detalle"])

    def test_isolated_saturday_lunch_return_reference_punch_is_not_left_blank(self) -> None:
        result = clasificar_checadas(
            ["12:29:58"],
            SATURDAY_SHIFT,
            estados=["Estado invalido"],
            dispositivos=["CHECADOR_PRODUCCION"],
        )
        self.assertIsNone(result["entrada"])
        self.assertIsNone(result["inicio_comida"])
        self.assertEqual(result["fin_comida"], "12:29:58")
        self.assertIsNone(result["salida"])
        self.assertEqual(result["status"], "Incidencia")
        self.assertEqual(
            result["detalle"],
            "Sin entrada | Sin inicio de comida | Sin salida final",
        )
        self.assertNotIn("Registro ambiguo", result["detalle"])

    def test_late_entry_with_later_exit_has_no_false_lunch_excess(self) -> None:
        result = clasificar_checadas(
            ["11:00:00", "12:40:00", "17:10:00"],
            WEEKDAY_SHIFT,
        )
        self.assertEqual(result["entrada"], "11:00:00")
        self.assertIsNone(result["inicio_comida"])
        self.assertIsNone(result["fin_comida"])
        self.assertEqual(result["salida"], "17:10:00")
        self.assertEqual(
            result["detalle"],
            "Retardo grave (180 min) | Registro incompleto",
        )
        self.assertNotIn("Exceso de comida", result["detalle"])

    def test_1130_with_final_exit_is_late_entry_not_lunch(self) -> None:
        result = clasificar_checadas(["11:30:00", "17:00:00"], WEEKDAY_SHIFT)
        self.assertEqual(result["entrada"], "11:30:00")
        self.assertEqual(result["salida"], "17:00:00")
        self.assertNotIn("Exceso de comida", result["detalle"])

    def test_single_unclear_punch_retains_ambiguity(self) -> None:
        result = clasificar_checadas(["11:30:00"], WEEKDAY_SHIFT)
        self.assertEqual(result["status"], "Ambiguo")
        self.assertIn("Registro ambiguo", result["detalle"])
        self.assertIn("Checada registrada sin clasificar (11:30:00)", result["detalle"])

    def test_single_saturday_entry_hint_before_lunch_is_shown_as_late_entry(self) -> None:
        result = clasificar_checadas(
            ["10:38:03"],
            SATURDAY_SHIFT,
            estados=["Entrada"],
            dispositivos=["CHECADOR_PRODUCCION"],
        )
        self.assertEqual(result["entrada"], "10:38:03")
        self.assertIsNone(result["inicio_comida"])
        self.assertIsNone(result["fin_comida"])
        self.assertIsNone(result["salida"])
        self.assertEqual(result["status"], "Retardo + incidencia")
        self.assertEqual(
            result["detalle"],
            "Retardo grave (158 min) | Sin inicio de comida | Sin regreso de comida | Sin salida final",
        )
        self.assertNotIn("Registro ambiguo", result["detalle"])

    def test_late_entry_with_complete_lunch_without_final_exit_is_not_ambiguous(self) -> None:
        result = clasificar_checadas(
            ["11:59:00", "12:09:00", "12:48:00"],
            WEEKDAY_SHIFT,
        )
        self.assertEqual(result["entrada"], "11:59:00")
        self.assertEqual(result["inicio_comida"], "12:09:00")
        self.assertEqual(result["fin_comida"], "12:48:00")
        self.assertIsNone(result["salida"])
        self.assertEqual(result["status"], "Retardo + incidencia")
        self.assertEqual(result["detalle"], "Retardo grave (239 min) | Sin salida final")
        self.assertNotIn("Sin entrada", result["detalle"])
        self.assertNotIn("Registro ambiguo", result["detalle"])

    def test_declared_entry_and_exit_are_assigned_directly(self) -> None:
        result = clasificar_checadas(
            ["08:48:40", "14:25:55"],
            WEEKDAY_SHIFT,
            estados=["Entrada", "Salida"],
            dispositivos=["CHECADOR_OFICINA", "CHECADOR_OFICINA"],
        )
        self.assertEqual(result["entrada"], "08:48:40")
        self.assertIsNone(result["inicio_comida"])
        self.assertIsNone(result["fin_comida"])
        self.assertEqual(result["salida"], "14:25:55")
        self.assertEqual(
            result["detalle"],
            "Retardo (48 min) | Sin inicio de comida | Sin regreso de comida | Salida anticipada (154 min)",
        )
        self.assertNotIn("Registro ambiguo", result["detalle"])
        self.assertNotIn("Checada registrada sin clasificar", result["detalle"])

    def test_declared_entry_and_exit_still_calculate_worked_hours_without_lunch_pair(self) -> None:
        work_date = date(2026, 6, 5)
        events = pd.DataFrame(
            [
                {"id_usuario": "1", "tiempo": pd.Timestamp(f"{work_date} 08:00:30"), "estado": "Entrada"},
                {"id_usuario": "1", "tiempo": pd.Timestamp(f"{work_date} 17:00:00"), "estado": "Salida"},
            ]
        )
        row = analyze_operational_day(
            prepared_personal().iloc[[0]].copy(),
            events,
            work_date,
            schedule_for_date(work_date, []),
        ).iloc[0]
        self.assertEqual(row["Horas trabajadas"], "09:00")
        self.assertEqual(row["Detalle"], "Sin inicio de comida | Sin regreso de comida")

    def test_operational_detail_does_not_show_state_or_device_audit_context(self) -> None:
        work_date = date(2026, 6, 5)
        events = pd.DataFrame(
            [
                {
                    "id_usuario": "1",
                    "tiempo": pd.Timestamp(f"{work_date} {checked_at}"),
                    "estado": state,
                    "dispositivo": "CHECADOR_PRODUCCION",
                }
                for checked_at, state in (
                    ("11:59:00", "Entrada"),
                    ("12:09:00", "Salida a descanso"),
                    ("12:48:00", "Regreso descanso"),
                )
            ]
        )
        row = analyze_operational_day(
            prepared_personal().iloc[[0]].copy(),
            events,
            work_date,
            schedule_for_date(work_date, []),
        ).iloc[0]
        audit_column = next(column for column in row.index if str(column).startswith("Auditor"))
        self.assertEqual(row["Detalle"], "Retardo grave (239 min) | Sin salida final")
        self.assertNotIn("CHECADOR", row["Detalle"])
        self.assertNotIn("Salida a descanso", row["Detalle"])
        self.assertIn("dispositivo=CHECADOR_PRODUCCION", row[audit_column])
        self.assertIn("estado=Salida a descanso", row[audit_column])

    def test_single_punch_before_lunch_without_later_evidence_does_not_invent_entry(self) -> None:
        result = clasificar_checadas(["11:59:00"], WEEKDAY_SHIFT)
        self.assertIsNone(result["entrada"])
        self.assertEqual(result["inicio_comida"], "11:59:00")
        self.assertEqual(result["status"], "Incidencia")
        self.assertIn("Sin entrada", result["detalle"])

    def test_operational_day_shows_single_saturday_entry_punch_in_entry_column(self) -> None:
        work_date = date(2026, 6, 13)
        personal = pd.DataFrame([{"id_usuario": "39", "nombre_completo": "PERSONA TREINTA Y NUEVE"}])
        events = pd.DataFrame(
            [
                {
                    "id_usuario": "39",
                    "tiempo": pd.Timestamp(f"{work_date} 10:38:03"),
                    "estado_normalizado": "Entrada",
                    "dispositivo_normalizado": "CHECADOR_PRODUCCION",
                }
            ]
        )
        row = analyze_operational_day(
            personal,
            events,
            work_date,
            schedule_for_date(work_date, []),
        ).iloc[0]
        self.assertEqual(row["Entrada"], "10:38:03")
        self.assertEqual(
            row["Detalle"],
            "Retardo grave (158 min) | Sin inicio de comida | Sin regreso de comida | Sin salida final",
        )
        self.assertEqual(row["Estatus"], "Retardo + incidencia")


class DuplicatePunchNormalizationTests(unittest.TestCase):
    def test_nearby_entry_duplicate_does_not_create_false_lunch_excess(self) -> None:
        result = clasificar_checadas(
            ["08:03:34", "08:06:48", "12:03:12", "19:04:21"],
            WEEKDAY_SHIFT,
        )
        self.assertEqual(result["entrada"], "08:03:34")
        self.assertEqual(result["inicio_comida"], "12:03:12")
        self.assertIsNone(result["fin_comida"])
        self.assertEqual(result["salida"], "19:04:21")
        self.assertEqual(result["status"], "Retardo + incidencia")
        self.assertEqual(result["detalle"], "Retardo (3 min) | Sin regreso de comida")
        self.assertNotIn("Exceso de comida", result["detalle"])
        self.assertNotIn("Checada no reconocida", result["detalle"])
        self.assertIn("08:06:48", result["checadas_no_utilizadas"])
        self.assertEqual(
            result["checadas_duplicadas"],
            [
                {
                    "duplicada": "08:06:48",
                    "original": "08:03:34",
                    "diferencia_segundos": 194,
                    "bloque_probable": "entrada",
                }
            ],
        )
        self.assertIn("08:06:48 -> duplicada/no utilizada", result["auditoria"])
        self.assertIn("Se conserv", result["auditoria"])
        self.assertIn("08:03:34", result["auditoria"])

    def test_nearby_lunch_out_duplicate_is_not_used_as_lunch_return(self) -> None:
        result = clasificar_checadas(
            ["08:00:00", "12:00:05", "12:02:00", "12:30:00", "17:00:00"],
            WEEKDAY_SHIFT,
        )
        self.assertEqual(result["inicio_comida"], "12:00:05")
        self.assertEqual(result["fin_comida"], "12:30:00")
        self.assertEqual(result["salida"], "17:00:00")
        self.assertEqual(result["status"], "Puntual")
        self.assertEqual(result["detalle"], "")
        self.assertEqual(result["checadas_duplicadas"][0]["duplicada"], "12:02:00")

    def test_nearby_exit_duplicate_is_not_operational_incident(self) -> None:
        result = clasificar_checadas(
            ["08:00:00", "12:00:00", "12:45:00", "17:00:00", "17:03:00"],
            WEEKDAY_SHIFT,
        )
        self.assertEqual(result["salida"], "17:00:00")
        self.assertEqual(result["status"], "Puntual")
        self.assertEqual(result["detalle"], "")
        self.assertEqual(result["checadas_duplicadas"][0]["duplicada"], "17:03:00")

    def test_real_saturday_events_are_not_removed(self) -> None:
        result = clasificar_checadas(
            ["08:00:00", "12:00:00", "12:30:00", "14:00:00"],
            SATURDAY_SHIFT,
        )
        self.assertEqual(result["entrada"], "08:00:00")
        self.assertEqual(result["inicio_comida"], "12:00:00")
        self.assertEqual(result["fin_comida"], "12:30:00")
        self.assertEqual(result["salida"], "14:00:00")
        self.assertEqual(result["status"], "Puntual")
        self.assertEqual(result["checadas_duplicadas"], [])

    def test_operational_report_keeps_duplicates_in_audit_only(self) -> None:
        row = operational_row(
            ["08:03:34", "08:06:48", "12:03:12", "19:04:21"],
            work_date=date(2026, 6, 5),
        )
        self.assertEqual(row["Entrada"], "08:03:34")
        self.assertEqual(row["Inicio comida"], "12:03:12")
        self.assertEqual(row["Fin comida"], "")
        self.assertEqual(row["Salida"], "19:04:21")
        self.assertEqual(row["Detalle"], "Retardo (3 min) | Sin regreso de comida")
        self.assertNotIn("Exceso de comida", row["Detalle"])
        self.assertNotIn("Checada no reconocida", row["Detalle"])
        self.assertIn("08:06:48 -> duplicada/no utilizada", row["Auditoría clasificación"])


class OperationalEvaluationTests(unittest.TestCase):
    def test_schedule_selects_weekday_and_saturday_limits(self) -> None:
        weekday = schedule_for_date(date(2026, 6, 5), [])
        saturday = schedule_for_date(date(2026, 6, 6), [])
        self.assertTrue(weekday.is_workday)
        self.assertTrue(saturday.is_workday)
        self.assertEqual(weekday.lunch_max_minutes, 45)
        self.assertEqual(saturday.lunch_max_minutes, 30)

    def test_saturday_operational_evaluation_uses_30_minutes(self) -> None:
        exact = operational_row(
            ["08:00:00", "12:40:00", "13:10:00", "14:00:00"],
            work_date=date(2026, 6, 6),
        )
        over = operational_row(
            ["08:00:00", "12:40:00", "13:10:01", "14:00:00"],
            work_date=date(2026, 6, 6),
        )
        self.assertEqual(exact["Estatus"], "Puntual")
        self.assertEqual(over["Detalle"], "Exceso de comida (+1 min)")

    def test_partial_cutoff_does_not_report_lunch_omissions(self) -> None:
        row = operational_row(
            ["08:00:00"],
            cutoff_time=datetime(2026, 6, 5, 13, 30),
        )
        self.assertEqual(row["Estatus"], "Puntual")
        self.assertEqual(row["Detalle"], "")

    def test_worked_hours_require_entry_and_exit_and_reject_incomplete_lunch(self) -> None:
        scheduled_entry = datetime(2026, 6, 5, 8, 0)
        entry = datetime(2026, 6, 5, 8, 0)
        lunch_out = datetime(2026, 6, 5, 12, 0)
        lunch_return = datetime(2026, 6, 5, 12, 45)
        exit_time = datetime(2026, 6, 5, 17, 0)
        self.assertEqual(
            calculate_worked_minutes(entry, None, None, exit_time, scheduled_entry=scheduled_entry),
            540,
        )
        self.assertIsNone(calculate_worked_minutes(entry, lunch_out, None, exit_time, scheduled_entry=scheduled_entry))
        self.assertIsNone(calculate_worked_minutes(entry, None, lunch_return, exit_time, scheduled_entry=scheduled_entry))
        self.assertEqual(
            calculate_worked_minutes(entry, lunch_out, lunch_return, exit_time, scheduled_entry=scheduled_entry),
            495,
        )

    def test_worked_hours_use_scheduled_entry_until_08_00_59(self) -> None:
        scheduled_entry = datetime(2026, 6, 5, 8, 0)
        self.assertEqual(
            calculate_worked_minutes(
                datetime(2026, 6, 5, 8, 0, 59),
                None,
                None,
                datetime(2026, 6, 5, 17, 0),
                scheduled_entry=scheduled_entry,
            ),
            540,
        )
        self.assertEqual(
            calculate_worked_minutes(
                datetime(2026, 6, 5, 8, 1, 0),
                None,
                None,
                datetime(2026, 6, 5, 17, 0),
                scheduled_entry=scheduled_entry,
            ),
            539,
        )

    def test_isolated_lunch_time_punch_has_no_worked_hours(self) -> None:
        row = operational_row(["12:01:00"])
        self.assertEqual(row["Horas trabajadas"], "")


class SundayNonWorkingDayTests(unittest.TestCase):
    def test_sunday_has_no_fallback_schedule(self) -> None:
        schedule = schedule_for_date(date(2026, 6, 7), [])
        self.assertFalse(schedule.is_workday)
        self.assertEqual(schedule.label, "Domingo - día no laborable")
        self.assertIsNone(schedule.entry_time)
        self.assertIsNone(schedule.lunch_max_minutes)

    def test_sunday_punches_are_review_only(self) -> None:
        work_date = date(2026, 6, 7)
        rows = analyze_operational_day(
            prepared_personal(),
            prepared_events(work_date, {"1": ("09:10:00", "13:25:00")}),
            work_date,
            schedule_for_date(work_date, []),
        )
        reviewed = rows.loc[rows["ID"] == "1"].iloc[0]
        neutral = rows.loc[rows["ID"] == "2"].iloc[0]
        self.assertEqual(reviewed["Estatus"], "Revisión")
        self.assertEqual(reviewed["Detalle"], "Checadas en día no laborable")
        self.assertEqual(reviewed["Entrada"], "")
        self.assertEqual(reviewed["Horas trabajadas"], "")
        self.assertEqual(reviewed["Retardo min"], 0)
        self.assertIn("09:10:00, 13:25:00", reviewed["Auditoría clasificación"])
        self.assertEqual(neutral["Estatus"], "Día no laborable")
        self.assertEqual(neutral["Detalle"], "")

    def test_daily_sunday_metrics_remain_zero(self) -> None:
        with TemporaryDirectory() as tmp:
            root = Path(tmp)
            personal, events = write_sources(
                root,
                {
                    "1": ("PERSONA", "UNO", ("09:10:00",)),
                    "2": ("PERSONA", "DOS", ()),
                },
                work_date=date(2026, 6, 7),
            )
            result = calculate_attendance(personal, events, root / "out")
            self.assertEqual(result.schedule_label, "Domingo - día no laborable")
            self.assertEqual(result.attendance_count, 0)
            self.assertEqual(result.tardy_count, 0)
            self.assertEqual(result.absence_count, 0)
            self.assertEqual(result.incident_employee_count, 0)
            self.assertTrue(result.tardy_frame.empty)
            self.assertTrue(result.absence_frame.empty)
            self.assertTrue(result.incident_frame.empty)
            self.assertTrue(any("domingo no laborable" in issue.message for issue in result.issues))

    def test_sunday_only_range_generates_review_without_workday_metrics(self) -> None:
        with TemporaryDirectory() as tmp:
            root = Path(tmp)
            personal, events = write_sources(
                root,
                {"1": ("PERSONA", "UNO", ("09:10:00", "13:25:00"))},
                work_date=date(2026, 6, 7),
            )
            result = calculate_attendance_range(personal, events, root / "out")
            self.assertEqual(result.workday_count, 0)
            self.assertEqual(result.operational_day_count, 0)
            self.assertEqual(result.attendance_count, 0)
            self.assertEqual(result.tardy_count, 0)
            self.assertEqual(result.absence_count, 0)
            self.assertEqual(result.incident_employee_count, 0)
            self.assertEqual(result.detail_frame.iloc[0]["Estatus"], "Revisión")

    def test_range_does_not_count_sunday_review_as_attendance(self) -> None:
        with TemporaryDirectory() as tmp:
            root = Path(tmp)
            personal_path = root / "personal.xlsx"
            events_path = root / "range.xlsx"
            pd.DataFrame(
                [{"ID de usuario": "1", "Nombre": "PERSONA", "Apellido": "UNO"}]
            ).to_excel(personal_path, sheet_name="data", index=False)
            rows = [
                {
                    "Tiempo": f"{day} {checked_at}",
                    "ID de usuario": "1",
                    "Nombre": "PERSONA",
                    "Apellido": "UNO",
                    "Estado": "",
                }
                for day, times in {
                    "2026-06-05": ("08:00:00", "12:00:00", "12:45:00", "17:00:00"),
                    "2026-06-07": ("09:10:00",),
                }.items()
                for checked_at in times
            ]
            pd.DataFrame(rows).to_excel(events_path, sheet_name="data", index=False)
            result = calculate_attendance_range(personal_path, events_path, root / "out")
            self.assertEqual(result.workday_count, 2)
            self.assertEqual(result.operational_day_count, 1)
            self.assertEqual(result.non_operational_day_count, 1)
            self.assertEqual(result.attendance_count, 1)
            self.assertEqual(result.incident_employee_count, 0)
            sunday = result.detail_frame[result.detail_frame["Fecha"] == "2026-06-07"].iloc[0]
            self.assertEqual(sunday["Estatus"], "Revisión")


class ReportingAndDocumentationTests(unittest.TestCase):
    def test_configuration_supports_schedule_and_employee_overrides(self) -> None:
        with TemporaryDirectory() as tmp:
            path = Path(tmp) / "classification.json"
            path.write_text(
                """
                {
                  "turnos": {"Sábado": {"salida": {"despues": 300, "max_despues": 360}}},
                  "empleados": {"14": {"entrada": {"despues": 180, "max_despues": 240}}}
                }
                """,
                encoding="utf-8",
            )
            config = load_classification_configuration(path)
            self.assertIn("Sábado", config.schedule_policies)
            self.assertIn("14", config.employee_policies)

    def test_quick_view_rejects_personnel_not_in_original_source(self) -> None:
        authorized = pd.DataFrame(
            [{"id_usuario": "1", "nombre_completo": "EMPLEADO FUENTE"}]
        )
        daily = pd.DataFrame(
            [
                {"ID": "1", "Nombre": "EMPLEADO FUENTE"},
                {"ID": "20", "Nombre": "REGISTRO NO AUTORIZADO"},
            ]
        )
        with self.assertRaisesRegex(ValueError, "no proviene de la BBDD"):
            build_quick_view_source(daily, authorized)

    def test_operational_report_rejects_fixture_source_paths(self) -> None:
        with TemporaryDirectory() as tmp:
            root = Path(tmp)
            fixture_dir = root / "tests" / "fixtures"
            fixture_dir.mkdir(parents=True)
            personal = fixture_dir / "personal.xlsx"
            events = fixture_dir / "events.xlsx"
            personal.touch()
            events.touch()
            with self.assertRaisesRegex(ValueError, "reporte operativo"):
                calculate_attendance(personal, events, root / "out")

    def test_operational_report_rejects_demo_output_paths(self) -> None:
        with TemporaryDirectory() as tmp:
            root = Path(tmp)
            personal = root / "personal.xlsx"
            events = root / "events.xlsx"
            personal.touch()
            events.touch()
            with self.assertRaisesRegex(ValueError, "reporte operativo"):
                calculate_attendance(personal, events, root / "demo" / "out")

    def test_unknown_event_names_do_not_enter_operational_report(self) -> None:
        with TemporaryDirectory() as tmp:
            root = Path(tmp)
            personal_path = root / "personal.xlsx"
            events_path = root / "events.xlsx"
            pd.DataFrame(
                [{"ID de usuario": "1", "Nombre": "EMPLEADO", "Apellido": "FUENTE"}]
            ).to_excel(personal_path, sheet_name="data", index=False)
            pd.DataFrame(
                [
                    {
                        "Tiempo": f"2026-06-05 {checked_at}",
                        "ID de usuario": "1",
                        "Nombre": "EMPLEADO",
                        "Apellido": "FUENTE",
                        "Estado": "",
                    }
                    for checked_at in ("08:00:00", "12:00:00", "12:45:00", "17:00:00")
                ]
                + [
                    {
                        "Tiempo": "2026-06-05 15:00:00",
                        "ID de usuario": "20",
                        "Nombre": "REGISTRO",
                        "Apellido": "NO AUTORIZADO",
                        "Estado": "",
                    }
                ]
            ).to_excel(events_path, sheet_name="data", index=False)
            result = calculate_attendance(personal_path, events_path, root / "out")
            workbook = load_workbook(result.report_file, data_only=True)
            try:
                text = " ".join(
                    str(cell.value)
                    for sheet in workbook.worksheets
                    for row in sheet.iter_rows()
                    for cell in row
                    if cell.value is not None
                )
                self.assertIn("EMPLEADO FUENTE", text)
                self.assertNotIn("REGISTRO NO AUTORIZADO", text)
            finally:
                workbook.close()

    def test_report_contains_no_overtime_and_keeps_audit_separate(self) -> None:
        with TemporaryDirectory() as tmp:
            root = Path(tmp)
            personal, events = write_sources(
                root,
                {"1": ("EMPLEADO", "AISLADO", ("12:01:00",))},
                work_date=date(2026, 6, 5),
            )
            result = calculate_attendance(personal, events, root / "out")
            workbook = load_workbook(result.report_file, data_only=True)
            try:
                text = " ".join(
                    str(cell.value)
                    for sheet in workbook.worksheets
                    for row in sheet.iter_rows()
                    for cell in row
                    if cell.value is not None
                )
                self.assertNotIn(REMOVED_VISIBLE_LABEL, text)
                self.assertIn("Auditoría clasificación", workbook.sheetnames)
                self.assertIn("Sin entrada | Sin regreso de comida | Sin salida final", text)
            finally:
                workbook.close()

    def test_sunday_report_keeps_raw_times_only_in_audit(self) -> None:
        with TemporaryDirectory() as tmp:
            root = Path(tmp)
            personal, events = write_sources(
                root,
                {"1": ("PERSONA", "UNO", ("09:10:00", "09:11:00"))},
                work_date=date(2026, 6, 7),
            )
            result = calculate_attendance(personal, events, root / "out")
            row = result.daily_frame.iloc[0]
            self.assertEqual(row["Entrada"], "")
            self.assertEqual(row["Inicio comida"], "")
            self.assertIn("09:10:00, 09:11:00", row["Auditoría clasificación"])
            workbook = load_workbook(result.report_file, data_only=True)
            try:
                detail_text = " ".join(
                    str(cell.value)
                    for row in workbook["Detalle diario"].iter_rows()
                    for cell in row
                    if cell.value is not None
                )
                audit_text = " ".join(
                    str(cell.value)
                    for row in workbook["Auditoría clasificación"].iter_rows()
                    for cell in row
                    if cell.value is not None
                )
                self.assertNotIn("09:10:00", detail_text)
                self.assertNotIn("09:11:00", detail_text)
                self.assertIn("09:10:00", audit_text)
                self.assertIn("09:11:00", audit_text)
            finally:
                workbook.close()

    def test_range_report_keeps_visible_contract_without_overtime(self) -> None:
        with TemporaryDirectory() as tmp:
            root = Path(tmp)
            personal, events = write_sources(
                root,
                {"1": ("EMPLEADO", "RETARDO", ("10:00:00", "13:00:00", "13:30:00", "17:00:00"))},
                work_date=date(2026, 6, 5),
            )
            result = calculate_attendance_range(personal, events, root / "out")
            workbook = load_workbook(result.report_file, data_only=True)
            try:
                text = " ".join(
                    str(cell.value)
                    for sheet in workbook.worksheets
                    for row in sheet.iter_rows()
                    for cell in row
                    if cell.value is not None
                )
                self.assertNotIn(REMOVED_VISIBLE_LABEL, text)
                self.assertIn("Retardo grave (120 min)", text)
            finally:
                workbook.close()

    def test_documentation_matches_current_business_behavior(self) -> None:
        root = Path(__file__).resolve().parents[1]
        readme = (root / "README.md").read_text(encoding="utf-8")
        changelog = (root / "CHANGELOG.md").read_text(encoding="utf-8")
        docs = (root / "docs" / "clasificacion_contextual.md").read_text(encoding="utf-8")
        for text in (readme, docs):
            self.assertIn("Día no laborable", text)
            self.assertIn("Revisión", text)
            self.assertIn("30:01", text)
            self.assertIn("45:01", text)
        self.assertIn("Sunday is treated as a non-working day", changelog)
        self.assertNotIn(REMOVED_VISIBLE_LABEL.lower(), readme.lower())
        removed_section = changelog.split("### Removed", 1)[1].split("## ", 1)[0].strip()
        self.assertEqual(
            removed_section,
            "- Overtime calculation and overtime report generation.",
        )


if __name__ == "__main__":
    unittest.main()
